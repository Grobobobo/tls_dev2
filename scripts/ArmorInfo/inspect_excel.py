#!/usr/bin/env python3
import openpyxl
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

from pathlib import Path
SCRIPT_DIR = Path(__file__).resolve().parent
wb = openpyxl.load_workbook(SCRIPT_DIR / 'TLS_ITEM_VALUES.xlsx', data_only=True)

print("Sheets:", wb.sheetnames)

for sheet_name in ['New Body Armors', 'New Helmets', 'New Pants']:
    if sheet_name not in wb.sheetnames:
        print(f"\nSheet '{sheet_name}' NOT FOUND")
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print("\nFirst 5 rows:")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)
    print("\nAll non-empty rows (first 100):")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        if any(v is not None for v in row):
            print(f"  Row {i}: {row}")
