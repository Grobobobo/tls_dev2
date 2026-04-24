#!/usr/bin/env python3
import openpyxl
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

from pathlib import Path
SCRIPT_DIR = Path(__file__).resolve().parent
wb = openpyxl.load_workbook(SCRIPT_DIR / 'TLS_ITEM_VALUES.xlsx', data_only=True)

for sheet_name in ['Vanilla Body Armors', 'Vanilla Helmets', 'Vanilla Pants']:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print("All non-empty rows:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        if any(v is not None for v in row):
            print(f"  Row {i}: {row}")
