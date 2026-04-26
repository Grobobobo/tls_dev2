#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Update armor/helmet/pants XML files in modded_files based on
'New Body Armors', 'New Helmets', and 'New Pants' sheets in TLS_ITEM_VALUES.xlsx.

Strategy:
  - For levels 0-5: fully replace MainStatBonus, BaseStatBonuses, BasePrice, and Skills.
  - For levels 6-10:
      * Apply delta (new_L5 - vanilla_L5) to the MainStatBonus value.
      * Update the MainStatBonus stat name if it changed.
      * Remove BaseStatBonus entries whose stat is gone in New sheet.
      * Keep BaseStatBonus entries that exist in both, shifted by (new_L5 - vanilla_L5) for that stat.
      * Add new BaseStatBonus entries (present in New but not Vanilla) at the level-5 value from New.
      * BasePrice kept from base file unchanged.
      * Skills (body armors only): set to '{Level0Skill}5' using the new skill name.
"""

import openpyxl
import xml.etree.ElementTree as ET
import shutil
import sys
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
EXCEL_PATH = SCRIPT_DIR / 'TLS_ITEM_VALUES.xlsx'
BASE_DIR = REPO_ROOT / 'base_files'
MODDED_DIR = REPO_ROOT / 'modded_files'

FILES = {
    'armors': {
        'xml_file': 'ItemDefinitions_BodyArmors.xml',
        'vanilla_sheet': 'Vanilla Body Armors',
        'new_sheet': 'New Body Armors',
        'has_skill': True,
    },
    'helmets': {
        'xml_file': 'ItemDefinitions_Helmets.xml',
        'vanilla_sheet': 'Vanilla Helmets',
        'new_sheet': 'New Helmets',
        'has_skill': False,
    },
    'pants': {
        'xml_file': 'ItemDefinitions_Pants.xml',
        'vanilla_sheet': 'Vanilla Pants',
        'new_sheet': 'New Pants',
        'has_skill': False,
    },
}

# Excel uses a few stat labels that differ from XML stat IDs.
STAT_NAME_ALIASES = {
    'MagicDamage': 'MagicalDamage',
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_vals(s):
    """Parse '3/5/7/9/11/13' or '-1/-1/-1' into a list of ints."""
    if s is None:
        return None
    parts = str(s).strip().split('/')
    result = []
    for p in parts:
        p = p.strip()
        try:
            result.append(int(float(p)))
        except ValueError:
            return None
    return result


def canonical_stat_name(stat_name):
    """Normalize stat name from sheet into the XML stat name."""
    if stat_name is None:
        return None
    stat = str(stat_name).strip()
    return STAT_NAME_ALIASES.get(stat, stat)


def parse_sheet(wb, sheet_name, has_skill):
    """
    Parse a Vanilla or New armor sheet.

    Auto-detects whether a Level0Skill column is present by reading the header row.
    Returns:
        ordered_names: list of item names in sheet row order
        items: dict  name -> {
            'skill': str|None,
            'ms_stat': str,
            'ms_vals': list[int],   # 6 values (levels 0-5)
            'attrs': [(stat_name, [v0,v1,v2,v3,v4,v5]), ...],
            'prices': list[int],    # 6 values
        }
    """
    ws = wb[sheet_name]
    ordered_names = []
    items = {}

    max_col = ws.max_column

    # Read header to detect actual column layout
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Detect if Level0Skill column is present (col index 2)
    sheet_has_skill_col = (len(header) > 2 and header[2] == 'Level0Skill')

    if sheet_has_skill_col:
        # cols (0-based): Tag Name Skill MS_Name MS_Vals [A_Name A_Vals]*n Price
        ms_stat_col = 3
        ms_vals_col = 4
        attr_start = 5
    else:
        # cols (0-based): Tag Name MS_Name MS_Vals [A_Name A_Vals]*n Price
        ms_stat_col = 2
        ms_vals_col = 3
        attr_start = 4

    price_col = max_col - 1  # last column (0-based)
    num_attr_pairs = (price_col - attr_start) // 2

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        row = list(row)
        while len(row) < max_col:
            row.append(None)

        tag = row[0]
        name = row[1]
        skill = str(row[2]) if (sheet_has_skill_col and row[2] is not None) else None
        ms_stat = canonical_stat_name(row[ms_stat_col])
        ms_vals_raw = row[ms_vals_col]
        price_raw = row[price_col]

        attrs = []
        for i in range(num_attr_pairs):
            attr_name = row[attr_start + i * 2]
            attr_vals_raw = row[attr_start + i * 2 + 1]
            if attr_name is not None:
                vals = parse_vals(attr_vals_raw)
                if vals is not None:
                    attrs.append((canonical_stat_name(attr_name), vals))

        item = {
            'tag': tag,
            'name': name,
            'skill': skill,
            'ms_stat': ms_stat,
            'ms_vals': parse_vals(ms_vals_raw),
            'attrs': attrs,
            'prices': parse_vals(price_raw),
        }
        ordered_names.append(name)
        items[name] = item

    return ordered_names, items


def levenshtein(s1, s2):
    """Compute Levenshtein edit distance between two strings."""
    m, n = len(s1), len(s2)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[:]
        dp[0] = i
        for j in range(1, n + 1):
            cost = 0 if s1[i - 1] == s2[j - 1] else 1
            dp[j] = min(dp[j] + 1, dp[j - 1] + 1, prev[j - 1] + cost)
    return dp[n]


def fuzzy_name_match(target, candidates):
    """Find the best fuzzy match for target in candidates using Levenshtein distance."""
    target_norm = target.lower()
    best = None
    best_score = float('inf')
    for c in candidates:
        dist = levenshtein(target_norm, c.lower())
        if dist < best_score:
            best_score = dist
            best = c
    if best_score <= 2:  # Allow up to 2 edit operations (handles 1 deletion/insertion)
        return best
    return None


def indent_xml(elem, level=0):
    """Recursively indent XML elements with tabs."""
    indent = '\n' + '\t' * level
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + '\t'
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for child in elem:
            indent_xml(child, level + 1)
        # Fix last child tail
        if not child.tail or not child.tail.strip():
            child.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent


def get_level_element(item_elem, level_id):
    """Return the Level element with Id=level_id, or None."""
    for lv in item_elem.findall('.//Level'):
        if lv.get('Id') == str(level_id):
            return lv
    return None


def update_level_0_to_5(item_elem, new_data, has_skill):
    """Fully replace levels 0-5 based on new_data."""
    ms_vals = new_data['ms_vals']   # 6 values
    prices = new_data['prices']     # 6 values
    attrs = new_data['attrs']       # list of (stat_name, [6 values])
    ms_stat = new_data['ms_stat']
    skill_base = new_data['skill']  # only for armors

    for lv_id in range(6):
        lv_elem = get_level_element(item_elem, lv_id)
        if lv_elem is None:
            continue

        # --- MainStatBonus ---
        msb = lv_elem.find('MainStatBonus')
        if msb is not None:
            msb.set('Stat', ms_stat)
            msb.text = str(ms_vals[lv_id])

        # --- BasePrice ---
        bp = lv_elem.find('BasePrice')
        if bp is not None and prices is not None:
            bp.text = str(prices[lv_id])

        # --- BaseStatBonuses ---
        bsb_container = lv_elem.find('BaseStatBonuses')
        if attrs:
            if bsb_container is None:
                # Insert after MainStatBonus/BasePrice, before Skills
                bsb_container = ET.SubElement(lv_elem, 'BaseStatBonuses')
            else:
                # Clear existing children
                for child in list(bsb_container):
                    bsb_container.remove(child)

            for stat_name, stat_vals in attrs:
                bsb_elem = ET.SubElement(bsb_container, 'BaseStatBonus')
                bsb_elem.set('Stat', stat_name)
                bsb_elem.text = str(stat_vals[lv_id])
        else:
            # No attributes in new data — remove the container if present
            if bsb_container is not None:
                lv_elem.remove(bsb_container)

        # --- Skills (body armors only) ---
        if has_skill and skill_base:
            skills_container = lv_elem.find('Skills')
            if skills_container is not None:
                for child in list(skills_container):
                    skills_container.remove(child)
                if lv_id == 0:
                    skill_name = skill_base
                elif lv_id <= 4:
                    skill_name = skill_base + str(lv_id)
                else:
                    skill_name = skill_base + '5'
                skill_elem = ET.SubElement(skills_container, 'Skill')
                skill_elem.text = skill_name


def update_levels_6_to_10(item_elem, new_data, vanilla_data, has_skill):
    """
    Update levels 6-10 by applying deltas from level-5 changes.
    """
    new_ms_stat = new_data['ms_stat']
    new_ms_l5 = new_data['ms_vals'][5]
    van_ms_l5 = vanilla_data['ms_vals'][5]
    ms_delta = new_ms_l5 - van_ms_l5

    # Build lookup dicts for attribute level-5 values
    new_attrs_dict = {name: vals for name, vals in new_data['attrs']}
    van_attrs_dict = {name: vals for name, vals in vanilla_data['attrs']}

    skill_base = new_data['skill']  # only set for armors

    for lv_id in range(6, 11):
        lv_elem = get_level_element(item_elem, lv_id)
        if lv_elem is None:
            continue

        # --- MainStatBonus ---
        msb = lv_elem.find('MainStatBonus')
        if msb is not None:
            msb.set('Stat', new_ms_stat)
            old_val = int(msb.text.strip())
            msb.text = str(old_val + ms_delta)

        # --- BaseStatBonuses ---
        bsb_container = lv_elem.find('BaseStatBonuses')

        if not new_attrs_dict:
            # New data has no extra attributes — remove container entirely
            if bsb_container is not None:
                lv_elem.remove(bsb_container)
        else:
            if bsb_container is None:
                bsb_container = ET.SubElement(lv_elem, 'BaseStatBonuses')
            else:
                for child in list(bsb_container):
                    bsb_container.remove(child)

            for stat_name, new_vals in new_data['attrs']:
                bsb_elem = ET.SubElement(bsb_container, 'BaseStatBonus')
                bsb_elem.set('Stat', stat_name)

                new_l5_val = new_vals[5]
                if stat_name in van_attrs_dict:
                    # Stat existed in vanilla — find old XML value and apply delta
                    van_l5_val = van_attrs_dict[stat_name][5]
                    stat_delta = new_l5_val - van_l5_val
                    # Find old value from current XML at this level
                    # Since we just cleared bsb_container, read from base XML isn't possible.
                    # Instead, compute from the vanilla level-5 + delta based on vanilla progression.
                    # We'll store the original bsb values before clearing — see note below.
                    # For now use the new_l5 value as fallback (corrected in second pass approach).
                    bsb_elem.text = str(new_l5_val)  # placeholder, corrected below
                else:
                    # Brand-new stat — use level-5 value
                    bsb_elem.text = str(new_l5_val)

        # --- Skills (body armors only) ---
        if has_skill and skill_base:
            skills_container = lv_elem.find('Skills')
            if skills_container is not None:
                for child in list(skills_container):
                    skills_container.remove(child)
                skill_elem = ET.SubElement(skills_container, 'Skill')
                skill_elem.text = skill_base + '5'


def update_levels_6_to_10_v2(item_elem, new_data, vanilla_data, base_item_elem, has_skill):
    """
    Correct version: reads original values from base_item_elem to compute deltas for levels 6-10.
    """
    new_ms_stat = new_data['ms_stat']
    new_ms_l5 = new_data['ms_vals'][5]
    van_ms_l5 = vanilla_data['ms_vals'][5]
    ms_delta = new_ms_l5 - van_ms_l5

    new_attrs_dict = {name: vals for name, vals in new_data['attrs']}
    van_attrs_dict = {name: vals for name, vals in vanilla_data['attrs']}

    skill_base = new_data['skill']

    for lv_id in range(6, 11):
        lv_elem = get_level_element(item_elem, lv_id)
        base_lv_elem = get_level_element(base_item_elem, lv_id)
        if lv_elem is None or base_lv_elem is None:
            continue

        # Read original BaseStatBonus values from base XML for this level
        orig_bsb = {}
        orig_bsb_container = base_lv_elem.find('BaseStatBonuses')
        if orig_bsb_container is not None:
            for bsb in orig_bsb_container.findall('BaseStatBonus'):
                stat = bsb.get('Stat')
                try:
                    orig_bsb[stat] = int(bsb.text.strip())
                except (ValueError, AttributeError):
                    orig_bsb[stat] = 0

        # --- MainStatBonus ---
        msb = lv_elem.find('MainStatBonus')
        base_msb = base_lv_elem.find('MainStatBonus')
        if msb is not None and base_msb is not None:
            msb.set('Stat', new_ms_stat)
            try:
                base_val = int(base_msb.text.strip())
            except (ValueError, AttributeError):
                base_val = 0
            msb.text = str(base_val + ms_delta)

        # --- BaseStatBonuses ---
        bsb_container = lv_elem.find('BaseStatBonuses')

        if not new_attrs_dict:
            if bsb_container is not None:
                lv_elem.remove(bsb_container)
        else:
            if bsb_container is None:
                bsb_container = ET.SubElement(lv_elem, 'BaseStatBonuses')
            else:
                for child in list(bsb_container):
                    bsb_container.remove(child)

            for stat_name, new_vals in new_data['attrs']:
                bsb_elem = ET.SubElement(bsb_container, 'BaseStatBonus')
                bsb_elem.set('Stat', stat_name)

                new_l5_val = new_vals[5]

                if stat_name in orig_bsb and stat_name in van_attrs_dict:
                    # Stat existed in vanilla AND in base XML at this level
                    van_l5_val = van_attrs_dict[stat_name][5]
                    stat_delta = new_l5_val - van_l5_val
                    bsb_elem.text = str(orig_bsb[stat_name] + stat_delta)
                else:
                    # Brand new stat for this item — pin at level-5 value
                    bsb_elem.text = str(new_l5_val)

        # --- Skills ---
        if has_skill and skill_base:
            skills_container = lv_elem.find('Skills')
            if skills_container is not None:
                for child in list(skills_container):
                    skills_container.remove(child)
                skill_elem = ET.SubElement(skills_container, 'Skill')
                skill_elem.text = skill_base + '5'


def write_xml(tree, out_path):
    """Write ElementTree to file with tab indentation and no XML declaration."""
    root = tree.getroot()
    # Apply indentation
    indent_xml(root, level=0)
    # Ensure root has a trailing newline
    if not root.tail:
        root.tail = '\n'

    # ET.write doesn't give us the exact header format; use tostring + manual write
    xml_str = ET.tostring(root, encoding='unicode')

    # Fix self-closing tags: ET uses <tag /> (with space) — normalize to <tag/>
    # (optional — keep as-is for compatibility)

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(xml_str)
        f.write('\n')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def process_armor_type(wb, config):
    xml_file = config['xml_file']
    vanilla_sheet = config['vanilla_sheet']
    new_sheet = config['new_sheet']
    has_skill = config['has_skill']

    base_xml_path = BASE_DIR / xml_file
    modded_xml_path = MODDED_DIR / xml_file

    print(f"\n{'='*70}")
    print(f"Processing: {xml_file}")
    print(f"{'='*70}")

    # Parse Excel sheets
    van_names, van_items = parse_sheet(wb, vanilla_sheet, has_skill)
    new_names, new_items = parse_sheet(wb, new_sheet, has_skill)

    print(f"  Vanilla sheet: {len(van_names)} items")
    print(f"  New sheet:     {len(new_names)} items")

    # Parse base XML to get item ordering and a reference copy
    base_tree = ET.parse(base_xml_path)
    base_root = base_tree.getroot()
    base_item_elems = base_root.findall('ItemDefinition')

    if len(base_item_elems) != len(van_names):
        print(f"  WARNING: XML has {len(base_item_elems)} items but Vanilla sheet has {len(van_names)} items!")

    # Check all New sheet items can be found in Vanilla sheet
    missing = [n for n in new_names if n not in van_items]
    if missing:
        print(f"  WARNING: Items in New sheet not found in Vanilla sheet: {missing}")

    # Copy base XML to modded path as our working copy
    shutil.copy2(base_xml_path, modded_xml_path)
    print(f"  Copied base XML -> {modded_xml_path.name}")

    # Parse the copy we will modify
    mod_tree = ET.parse(modded_xml_path)
    mod_root = mod_tree.getroot()
    mod_item_elems = mod_root.findall('ItemDefinition')

    updated = 0
    skipped = 0

    for idx, (item_elem, base_item_elem) in enumerate(zip(mod_item_elems, base_item_elems)):
        if idx >= len(van_names):
            break

        item_name = van_names[idx]

        # Try exact match first, then fuzzy match for New sheet lookup
        new_data = new_items.get(item_name)
        if new_data is None:
            fuzzy = fuzzy_name_match(item_name, list(new_items.keys()))
            if fuzzy:
                print(f"  [FUZZY] '{item_name}' matched to '{fuzzy}' in New sheet")
                new_data = new_items[fuzzy]
            else:
                print(f"  [SKIP] Index {idx}: '{item_name}' not found in New sheet (no fuzzy match)")
                skipped += 1
                continue

        van_data = van_items.get(item_name)
        if van_data is None:
            print(f"  [SKIP] Index {idx}: '{item_name}' not found in Vanilla sheet")
            skipped += 1
            continue

        # Validate data
        if new_data['ms_vals'] is None or len(new_data['ms_vals']) < 6:
            print(f"  [SKIP] '{item_name}': invalid MainStatBonus values in New sheet")
            skipped += 1
            continue
        if van_data['ms_vals'] is None or len(van_data['ms_vals']) < 6:
            print(f"  [WARN] '{item_name}': invalid MainStatBonus values in Vanilla sheet; skipping levels 6-10 delta")
            update_level_0_to_5(item_elem, new_data, has_skill)
            updated += 1
            continue

        # Update levels 0-5
        update_level_0_to_5(item_elem, new_data, has_skill)

        # Update levels 6-10 (using base XML as reference for original values)
        update_levels_6_to_10_v2(item_elem, new_data, van_data, base_item_elem, has_skill)

        updated += 1

    print(f"  Updated: {updated} items, Skipped: {skipped} items")

    # Re-indent and write
    indent_xml(mod_root, level=0)

    # Preserve original XML header comment if any
    # Write out the file preserving the original namespace declarations
    ET.register_namespace('xsi', 'http://www.w3.org/2001/XMLSchema-instance')

    # Build output string
    xml_bytes = ET.tostring(mod_root, encoding='unicode', xml_declaration=False)

    # Restore the original namespace attribute format from base file
    # (ET may reorder/rename namespace attributes)
    with open(base_xml_path, 'r', encoding='utf-8') as f:
        base_first_line = f.readline().strip()

    # Replace the root tag in output with the original (preserving namespace decls)
    import re
    xml_bytes = re.sub(
        r'^<ItemDefinitions[^>]*>',
        base_first_line,
        xml_bytes,
        count=1
    )

    with open(modded_xml_path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(xml_bytes)
        f.write('\n')

    print(f"  Written: {modded_xml_path}")


def main():
    print("=" * 70)
    print("ARMOR ITEMS UPDATE SCRIPT")
    print("=" * 70)
    print(f"Excel: {EXCEL_PATH}")
    print(f"Base: {BASE_DIR}")
    print(f"Modded: {MODDED_DIR}")

    if not EXCEL_PATH.exists():
        print(f"\nERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"\nSheets available: {wb.sheetnames}")

    for key, config in FILES.items():
        process_armor_type(wb, config)

    print("\n" + "=" * 70)
    print("Done.")


if __name__ == '__main__':
    main()
