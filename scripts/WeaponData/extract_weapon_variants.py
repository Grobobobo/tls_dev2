import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('tls_weapon_docs.xlsx', data_only=True)

# First, extract variant names from each weapon sheet (row 22)
print("=" * 80)
print("EXTRACTING VARIANT DEFINITIONS FROM WEAPON SHEETS")
print("=" * 80)

variant_stats = {}  # {weapon_name: {2: [stat1, stat2?], 3: [...], 4: [...], 5: [...]}}

weapon_sheets = [
    'sword', 'Hammer', '1h Axe', 'Dagger', '2h sword', '2H Hammer', '2H AXE', 'Spear',
    'Hand crossbow', 'Crossbow', 'Pistol', 'Shortbow', 'Longbow', 'Rifle',
    'Wand', 'Scepter', 'Tome of Secrets', 'Magic orb', 'power staff', 'druid staff',
    'War Shield', 'Claws', 'Cannon', 'Boomerang', 'Gauntlet', 'Sacred Flower'
]

for sheet_name in weapon_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"\n{sheet_name}: NOT FOUND")
        continue
    
    ws = wb[sheet_name]
    row_22 = {}
    
    # Row 22 has variant names for IDs ending 2-5 in columns A-D
    for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D'], start=2):
        cell = ws[f'{col_letter}22']
        value = cell.value
        if value:
            row_22[col_idx] = str(value)
    
    if row_22:
        variant_stats[sheet_name] = row_22
        print(f"\n{sheet_name}:")
        for variant_id, stats in sorted(row_22.items()):
            print(f"  ID {variant_id}: {stats}")

# Now extract Tier 1 and Tier 2 variant values
print("\n\n" + "=" * 80)
print("EXTRACTING TIER 1 VARIANT VALUES")
print("=" * 80)

tier1_sheet = wb['Tier 1 Variant Values']
tier1_data = {}

# Read header row (row 8)
headers = []
for col_idx in range(1, 20):  # Assuming up to 20 columns
    cell = tier1_sheet.cell(8, col_idx)
    if cell.value:
        headers.append(cell.value)
    else:
        break

print(f"Headers: {headers}")

# Read data rows (starting from row 9)
for row_idx in range(9, 25):  # Assuming up to 25 rows
    row_data = []
    has_data = False
    for col_idx in range(1, len(headers) + 1):
        cell = tier1_sheet.cell(row_idx, col_idx)
        if cell.value is not None:
            has_data = True
        row_data.append(cell.value)
    
    if has_data:
        tier1_data[row_idx - 8] = row_data
        print(f"Row {row_idx - 8}: {row_data}")

print("\n" + "=" * 80)
print("EXTRACTING TIER 2 VARIANT VALUES")
print("=" * 80)

tier2_sheet = wb['Tier 2 Variant Values']
tier2_data = {}

# Read header row (row 8)
headers2 = []
for col_idx in range(1, 20):
    cell = tier2_sheet.cell(8, col_idx)
    if cell.value:
        headers2.append(cell.value)
    else:
        break

print(f"Headers: {headers2}")

# Read data rows (starting from row 9)
for row_idx in range(9, 25):
    row_data = []
    has_data = False
    for col_idx in range(1, len(headers2) + 1):
        cell = tier2_sheet.cell(row_idx, col_idx)
        if cell.value is not None:
            has_data = True
        row_data.append(cell.value)
    
    if has_data:
        tier2_data[row_idx - 8] = row_data
        print(f"Row {row_idx - 8}: {row_data}")

# Save to JSON for further processing
output = {
    'variant_stats': variant_stats,
    'tier1_headers': headers,
    'tier1_data': tier1_data,
    'tier2_headers': headers2,
    'tier2_data': tier2_data
}

with open('weapon_variants.json', 'w') as f:
    json.dump(output, f, indent=2)

print("\n✓ Data saved to weapon_variants.json")
