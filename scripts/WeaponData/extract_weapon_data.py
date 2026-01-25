import openpyxl
import json

# Load with data_only to get calculated values
wb_data = openpyxl.load_workbook('tls_weapon_docs.xlsx', data_only=True)

sheets = wb_data.sheetnames
print(f"Found {len(sheets)} sheets\n")

weapon_data = {}

for sheet_name in sheets:
    ws = wb_data[sheet_name]
    
    # Row 5 has headers
    header_row = 5
    
    # Find column indices
    level_col = None
    min_dmg_col = None
    max_dmg_col = None
    
    for col_idx in range(1, 20):
        header_cell = ws.cell(header_row, col_idx).value
        if header_cell:
            val = str(header_cell).lower()
            if 'level' in val:
                level_col = col_idx
            elif 'new min damage' in val:
                min_dmg_col = col_idx
            elif 'new max damage' in val:
                max_dmg_col = col_idx
    
    if not (level_col and min_dmg_col and max_dmg_col):
        print(f"⚠️ {sheet_name}: Missing columns")
        continue
    
    # Extract damage values for levels (rows 6-12 for levels -1 to 5)
    levels = {}
    for row in range(6, 13):
        level_cell = ws.cell(row, level_col).value
        min_cell = ws.cell(row, min_dmg_col).value
        max_cell = ws.cell(row, max_dmg_col).value
        
        if level_cell is not None:
            try:
                level = int(level_cell) if isinstance(level_cell, int) else int(float(level_cell))
                min_dmg = int(min_cell) if isinstance(min_cell, int) else int(float(min_cell)) if min_cell else None
                max_dmg = int(max_cell) if isinstance(max_cell, int) else int(float(max_cell)) if max_cell else None
                if min_dmg is not None and max_dmg is not None:
                    levels[level] = {'min': min_dmg, 'max': max_dmg}
            except (ValueError, TypeError) as e:
                print(f"Error parsing {sheet_name} row {row}: {e}")
    
    weapon_data[sheet_name] = {
        'levels': levels
    }
    
    print(f"✓ {sheet_name}: {len(levels)} levels found")
    if levels:
        print(f"  Data: {json.dumps(levels)}")

# Save to a JSON file
with open('weapon_data.json', 'w') as f:
    json.dump(weapon_data, f, indent=2)

print("\n✓ Weapon data saved to weapon_data.json")
