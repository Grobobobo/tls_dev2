import json
from pathlib import Path
import xml.etree.ElementTree as ET

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
EXCEL_PATH = SCRIPT_DIR / 'tls_weapon_docs.xlsx'
SCROLL_XML_PATHS = [
    REPO_ROOT / 'modded_files' / 'ItemDefinitions_Usables',
    REPO_ROOT / 'modded_files' / 'ItemDefinitions_DLC1',
    REPO_ROOT / 'modded_files' / 'ItemDefinitions_DLC2',
]
CHANGE_LOG_PATH = SCRIPT_DIR / 'scroll_item_changes_corrected.json'
SKILL_DEFINITION_PATHS = [
    REPO_ROOT / 'modded_files' / 'SkillDefinitions_Items_Usables',
    REPO_ROOT / 'modded_files' / 'SkillDefinitions_DLC1',
    REPO_ROOT / 'modded_files' / 'SkillDefinitions_DLC2',
]


def load_weapon_data_from_excel(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    weapon_data = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        level_col = None
        min_damage_col = None
        max_damage_col = None

        for col_idx in range(1, 20):
            header_value = worksheet.cell(5, col_idx).value
            if not header_value:
                continue

            normalized_header = str(header_value).lower()
            if 'level' in normalized_header:
                level_col = col_idx
            elif 'new min damage' in normalized_header:
                min_damage_col = col_idx
            elif 'new max damage' in normalized_header:
                max_damage_col = col_idx

        if not (level_col and min_damage_col and max_damage_col):
            continue

        levels = {}
        for row_idx in range(6, 13):
            level_value = worksheet.cell(row_idx, level_col).value
            min_value = worksheet.cell(row_idx, min_damage_col).value
            max_value = worksheet.cell(row_idx, max_damage_col).value

            if level_value is None or min_value is None or max_value is None:
                continue

            try:
                level = int(float(level_value))
                min_damage = int(float(min_value))
                max_damage = int(float(max_value))
            except (TypeError, ValueError):
                continue

            levels[level] = {'min': min_damage, 'max': max_damage}

        if levels:
            weapon_data[sheet_name] = {'levels': levels}

    return weapon_data


def load_skill_definitions(file_paths):
    skill_definitions = {}

    for file_path in file_paths:
        tree = ET.parse(file_path)
        root = tree.getroot()

        for skill_def in root.findall('SkillDefinition'):
            skill_id = skill_def.get('Id')
            if skill_id:
                skill_definitions[skill_id] = skill_def

    return skill_definitions


def get_skill_damage_multiplier(skill_id, skill_definitions, visited=None):
    if not skill_id:
        return None

    if visited is None:
        visited = set()

    if skill_id in visited:
        return None

    visited.add(skill_id)
    skill_def = skill_definitions.get(skill_id)
    if skill_def is None:
        return None

    damage_multiplier = skill_def.find('./SkillAction/Attack/DamageMultiplier')
    if damage_multiplier is not None and damage_multiplier.text:
        try:
            return float(damage_multiplier.text)
        except ValueError:
            return None

    template_id = skill_def.get('TemplateId')
    if template_id:
        return get_skill_damage_multiplier(template_id, skill_definitions, visited)

    return None


weapon_data = load_weapon_data_from_excel(EXCEL_PATH)
skill_definitions = load_skill_definitions(SKILL_DEFINITION_PATHS)

# CORRECTED mapping of scroll item IDs to their source weapons
scroll_mapping = {
    'AxeBoomerangScroll': ('1h Axe', 'Axe'),
    'ThrowingDaggersScroll': ('Dagger', 'Dagger'),
    'ChargeScroll': ('2h sword', '2HSword'),              # FIXED: was Spear
    'SwordBlastScroll': ('2h sword', '2HSword'),          # FIXED: was sword
    'SuperSpinScroll': ('2H AXE', '2HAxe'),               # FIXED: was 2h sword
    'GroundSmashScroll': ('2H Hammer', '2HHammer'),
    'TripleSwipeScroll': ('Spear', 'Spear'),
    'GrapeshotScroll': ('Pistol', 'Pistol'),
    'RainOfArrowsScroll': ('Shortbow', 'Shortbow'),
    'ExplosiveBoltScroll': ('Crossbow', 'Crossbow'),
    'AssassinateScroll': ('Rifle', 'Rifle'),              # FIXED: was Dagger
    'MagicMissilesScroll': ('Wand', 'Wand'),
    'HammerOfFaithScroll': ('Scepter', 'Scepter'),
    'DeathRayScroll': ('Magic orb', 'Magic orb'),         # FIXED: was Tome of Secrets
    'ScorchingWaveScroll': ('power staff', 'power staff'), # FIXED: was Wand
    'FireThrowerScroll': ('power staff', 'power staff'),   # FIXED: was Rifle
    'FireballScroll': ('Tome of Secrets', 'Tome of Secrets'),  # FIXED: was Wand
    'LightningStrikeScroll': ('Tome of Secrets', 'Tome of Secrets'),  # FIXED: was Wand
    'BeeStingScroll': ('druid staff', 'druid staff'),      # FIXED: was Wand
    'MovingWallScroll': ('War Shield', 'WarShield'),
    'WallsOfPainScroll': ('War Shield', 'WarShield'),
    'ArmageddonScroll': ('Cannon', 'Cannon'),
    'GeoCannonScroll': ('Gauntlet', 'Gauntlet'),
    'PiercingClawsScroll': ('Claws', 'Claws'),
    'PainfulSpinScroll': ('Boomerang', 'Boomerang'),
    'PetalStormScroll': ('Sacred Flower', 'SacredFlower'),
    'TeleportationScroll': None,  # Should not have damage values
}

changes = []
total_updated = 0
total_removed = 0


def update_scroll_file(xml_path):
    global total_removed, total_updated

    tree = ET.parse(xml_path)
    root = tree.getroot()
    file_updated = 0
    file_removed = 0

    print(f"\nProcessing {xml_path.name}:")

    for item_def in root.findall('ItemDefinition'):
        item_id = item_def.get('Id')

        if item_id not in scroll_mapping:
            continue

        mapping = scroll_mapping[item_id]

        if mapping is None:
            print(f"\n  {item_id} (removing damage values):")
            level_variations = item_def.find('LevelVariations')
            if level_variations is not None:
                for level_elem in level_variations.findall('Level'):
                    base_damage = level_elem.find('BaseDamage')
                    if base_damage is not None:
                        level_id = level_elem.get('Id')
                        old_min = base_damage.get('Min')
                        old_max = base_damage.get('Max')
                        level_elem.remove(base_damage)
                        total_removed += 1
                        file_removed += 1
                        print(f"    Level {level_id}: Removed {old_min}-{old_max}")
            continue

        excel_sheet, _weapon_prefix = mapping

        if excel_sheet not in weapon_data:
            print(f"  Warning: {excel_sheet} not in weapon data")
            continue

        levels = weapon_data[excel_sheet]['levels']
        print(f"\n  {item_id} (from {excel_sheet}):")

        level_variations = item_def.find('LevelVariations')
        if level_variations is None:
            print("    No LevelVariations found")
            continue

        updated_count = 0
        for level_elem in level_variations.findall('Level'):
            level_id = level_elem.get('Id')
            try:
                excel_level = int(level_id)
            except (ValueError, TypeError):
                continue

            if excel_level not in levels:
                print(f"    Level {level_id}: Excel level {excel_level} not found")
                continue

            base_damage = level_elem.find('BaseDamage')
            if base_damage is None:
                continue

            old_min = base_damage.get('Min')
            old_max = base_damage.get('Max')
            new_min = levels[excel_level]['min']
            new_max = levels[excel_level]['max']

            skill_name = level_elem.findtext('./Skills/Skill')
            damage_multiplier = get_skill_damage_multiplier(skill_name, skill_definitions)
            if damage_multiplier is not None:
                new_min = round(new_min * damage_multiplier)
                new_max = round(new_max * damage_multiplier)

            if old_min != str(new_min) or old_max != str(new_max):
                base_damage.set('Min', str(new_min))
                base_damage.set('Max', str(new_max))
                changes.append({
                    'file': xml_path.name,
                    'scroll': item_id,
                    'level': level_id,
                    'excel_level': excel_level,
                    'old': f"{old_min}-{old_max}",
                    'new': f"{new_min}-{new_max}"
                })
                updated_count += 1
                file_updated += 1
                total_updated += 1
                print(f"    Level {level_id}: {old_min}-{old_max} -> {new_min}-{new_max}")

        if updated_count == 0:
            print("    No updates needed")

    tree.write(xml_path, encoding='utf-8', xml_declaration=True)
    return file_updated, file_removed


for scroll_xml_path in SCROLL_XML_PATHS:
    updated_count, removed_count = update_scroll_file(scroll_xml_path)
    print(f"  Summary: {updated_count} updated, {removed_count} removed")

print(f"\n✓ Updated {total_updated} scroll damage values")
print(f"✓ Removed {total_removed} damage values from TeleportationScroll")
print(f"✓ File saved")

with open(CHANGE_LOG_PATH, 'w', encoding='utf-8') as f:
    json.dump(changes, f, indent=2)
print(f"✓ Change log saved to scroll_item_changes_corrected.json")
