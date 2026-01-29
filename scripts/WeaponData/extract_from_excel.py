#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolidated weapon data update script.
Workflow:
  1. Extract weapon variant stats from Excel
  2. Extract weapon damage data from Excel
  3. Build stat bonus mapping (Excel names -> XML names)
  4. Update weapon stat bonuses in XML files
  5. Update scroll item damage values in XML files
"""

import openpyxl
import json
import os
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
import sys

# Fix encoding for Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

print("=" * 80)
print("WEAPON DATA CONSOLIDATION SCRIPT")
print("=" * 80)

# ============================================================================
# PHASE 1: EXTRACT WEAPON VARIANT STATS FROM EXCEL
# ============================================================================
print("\n[PHASE 1] Extracting weapon variant stats from Excel...")
print("-" * 80)

wb = openpyxl.load_workbook('tls_weapon_docs.xlsx', data_only=True)

# Extract variant names from each weapon sheet (row 22)
variant_stats = {}
weapon_sheets = [
    'sword', 'Hammer', '1h Axe', 'Dagger', '2h sword', '2H Hammer', '2H AXE', 'Spear',
    'Hand crossbow', 'Crossbow', 'Pistol', 'Shortbow', 'Longbow', 'Rifle',
    'Wand', 'Scepter', 'Tome of Secrets', 'Magic orb', 'power staff', 'druid staff',
    'War Shield', 'Claws', 'Cannon', 'Boomerang', 'Gauntlet', 'Sacred Flower'
]

for sheet_name in weapon_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  {sheet_name}: NOT FOUND")
        continue
    
    ws = wb[sheet_name]
    row_22 = {}
    
    # Row 22 has variant names for IDs ending 2-5 in columns A-D
    for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D'], start=2):
        cell = ws[f'{col_letter}22']
        value = cell.value
        if value:
            row_22[col_idx] = str(value)
    
    if row_22:
        variant_stats[sheet_name] = row_22
        print(f"✓ {sheet_name}: {len(row_22)} variants found")

# Extract Tier 1 Variant Values
tier1_sheet = wb['Tier 1 Variant Values']
tier1_headers = []
for col_idx in range(1, 20):
    cell = tier1_sheet.cell(8, col_idx)
    if cell.value:
        tier1_headers.append(cell.value)
    else:
        break

tier1_data = {}
for row_idx in range(9, 25):
    row_data = []
    has_data = False
    for col_idx in range(1, len(tier1_headers) + 1):
        cell = tier1_sheet.cell(row_idx, col_idx)
        if cell.value is not None:
            has_data = True
        row_data.append(cell.value)
    
    if has_data:
        tier1_data[row_idx - 8] = row_data

print(f"✓ Tier 1: {len(tier1_headers)} stat headers, {len(tier1_data)} data rows")

# Extract Tier 2 Variant Values
tier2_sheet = wb['Tier 2 Variant Values']
tier2_headers = []
for col_idx in range(1, 30):  # Increased to 30 to capture all headers
    cell = tier2_sheet.cell(8, col_idx)
    if cell.value:
        tier2_headers.append(cell.value)
    else:
        break

tier2_data = {}
for row_idx in range(9, 25):
    row_data = []
    has_data = False
    for col_idx in range(1, len(tier2_headers) + 1):  # Use full tier2_headers length
        cell = tier2_sheet.cell(row_idx, col_idx)
        if cell.value is not None:
            has_data = True
        row_data.append(cell.value)
    
    if has_data:
        tier2_data[row_idx - 8] = row_data

print(f"✓ Tier 2: {len(tier2_headers)} stat headers, {len(tier2_data)} data rows")

# ============================================================================
# PHASE 2: EXTRACT WEAPON DAMAGE DATA FROM EXCEL
# ============================================================================
print("\n[PHASE 2] Extracting weapon damage data from Excel...")
print("-" * 80)

wb_data = openpyxl.load_workbook('tls_weapon_docs.xlsx', data_only=True)
sheets = wb_data.sheetnames
weapon_data = {}

for sheet_name in sheets:
    ws = wb_data[sheet_name]
    
    # Row 5 has headers
    header_row = 5
    
    # Find column indices
    level_col = None
    min_dmg_col = None
    max_dmg_col = None
    
    for col_idx in range(1, 20):
        header_cell = ws.cell(header_row, col_idx).value
        if header_cell:
            val = str(header_cell).lower()
            if 'level' in val:
                level_col = col_idx
            elif 'new min damage' in val:
                min_dmg_col = col_idx
            elif 'new max damage' in val:
                max_dmg_col = col_idx
    
    if not (level_col and min_dmg_col and max_dmg_col):
        continue
    
    # Extract damage values for levels (rows 6-12 for levels -1 to 5)
    levels = {}
    for row in range(6, 13):
        level_cell = ws.cell(row, level_col).value
        min_cell = ws.cell(row, min_dmg_col).value
        max_cell = ws.cell(row, max_dmg_col).value
        
        if level_cell is not None:
            try:
                level = int(level_cell) if isinstance(level_cell, int) else int(float(level_cell))
                min_dmg = int(min_cell) if isinstance(min_cell, int) else int(float(min_cell)) if min_cell else None
                max_dmg = int(max_cell) if isinstance(max_cell, int) else int(float(max_cell)) if max_cell else None
                if min_dmg is not None and max_dmg is not None:
                    levels[level] = {'min': min_dmg, 'max': max_dmg}
            except (ValueError, TypeError):
                pass
    
    if levels:
        weapon_data[sheet_name] = {'levels': levels}

print(f"✓ Extracted damage data from {len(weapon_data)} weapon sheets")

# ============================================================================
# PHASE 3: BUILD STAT BONUS MAPPING
# ============================================================================
print("\n[PHASE 3] Building stat bonus mapping...")
print("-" * 80)

# Mapping from Excel stat names to internal XML stat names
stat_name_mapping = {
    # Main Stats
    'Momentum': 'MomentumAttacks',
    'Opportunism': 'OpportunisticAttacks',
    'Isolation': 'IsolatedAttacks',
    'Physical Damage': 'PhysicalDamage',
    'Ranged Damage': 'RangedDamage',
    'Magic Damage': 'MagicalDamage',
    'Skill Range': 'SkillRangeModifier',
    'Move Points': 'MovePointsTotal',
    'Dodge': 'Dodge',
    'Stun Chance': 'StunChanceModifier',
    'XP gain': 'ExperienceGainMultiplier',
    'Block': 'Block',
    'Health': 'HealthTotal',
    'Health Regen': 'HealthRegen',
    'Mana': 'ManaTotal',
    'Mana Regen': 'ManaRegen',
    'Reliability': 'Reliability',
    'Critical Power': 'CriticalPower',
    'Critical': 'Critical',
    'Poison Damage': 'PoisonDamageModifier',
    'Accuracy': 'Accuracy',
    'Armor': 'ArmorTotal',
    'Resistance': 'Resistance',
    'Resistance Reduction': 'ResistanceReduction',
    'Resistance reduction': 'ResistanceReduction',
    'Propagation Bounces': 'PropagationBouncesModifier',
    'Propagation Damage': 'PropagationDamage',
    
    # Composite stat keys
    'Health;Health Regen': ['HealthTotal', 'HealthRegen'],
    'Move Points;Dodge': ['MovePointsTotal', 'Dodge'],
    'Armor;Resistance': ['ArmorTotal', 'Resistance'],
    'Mana;Mana Regen': ['ManaTotal', 'ManaRegen'],
    
    # Case-insensitive variants
    'momentum': 'MomentumAttacks',
    'opportunism': 'OpportunisticAttacks',
    'isolation': 'IsolatedAttacks',
    'physical damage': 'PhysicalDamage',
    'ranged damage': 'RangedDamage',
    'magic damage': 'MagicalDamage',
    'skill range': 'SkillRangeModifier',
    'move points': 'MovePointsTotal',
    'dodge': 'Dodge',
    'stun chance': 'StunChanceModifier',
    'xp gain': 'ExperienceGainMultiplier',
    'block': 'Block',
    'health': 'HealthTotal',
    'health regen': 'HealthRegen',
    'mana': 'ManaTotal',
    'mana regen': 'ManaRegen',
    'reliability': 'Reliability',
    'critical power': 'CriticalPower',
    'critical': 'Critical',
    'poison damage': 'PoisonDamageModifier',
    'accuracy': 'Accuracy',
    'armor': 'ArmorTotal',
    'resistance': 'Resistance',
    'resistance reduction': 'ResistanceReduction',
    'propagation bounces': 'PropagationBouncesModifier',
    'propagation damage': 'PropagationDamage',
}

# Build Tier 1 bonuses dictionary
tier1_bonuses = {}
for level in range(6):
    level_key = level + 1  # tier1_data has integer keys 1-6
    tier1_bonuses[level] = {}
    if level_key in tier1_data:
        for header_idx, header in enumerate(tier1_headers):
            if header_idx < len(tier1_data[level_key]):
                try:
                    value = tier1_data[level_key][header_idx]
                    if isinstance(value, str):
                        try:
                            value = float(value)
                        except:
                            pass
                    else:
                        value = float(value)
                    tier1_bonuses[level][header] = value
                except (ValueError, TypeError):
                    pass

# Build Tier 2 bonuses dictionary
tier2_bonuses = {}
for level in range(6):
    level_key = level + 1  # tier2_data has integer keys 1-6
    tier2_bonuses[level] = {}
    if level_key in tier2_data:
        for header_idx, header in enumerate(tier2_headers):
            if header_idx < len(tier2_data[level_key]):
                try:
                    value = tier2_data[level_key][header_idx]
                    if isinstance(value, str):
                        try:
                            value = float(value)
                        except:
                            pass
                    else:
                        value = float(value)
                    tier2_bonuses[level][header] = value
                except (ValueError, TypeError):
                    pass

# Build weapon variants mapping
weapon_variants_mapping = {}
for weapon, variants in variant_stats.items():
    weapon_variants_mapping[weapon] = variants

print(f"✓ Stat name mapping: {len(stat_name_mapping)} entries")
print(f"✓ Tier 1 bonuses: {len(tier1_headers)} stats × 6 levels")
print(f"✓ Tier 2 bonuses: {len(tier2_headers)} stats × 6 levels")

# ============================================================================
# WEAPON CONSTANTS AND HELPER FUNCTIONS
# ============================================================================

# OffHand weapons to exclude
OFFHAND_WEAPONS = {'BattleMageMagicWand', 'BattleMageSword', 'DuelingPistol', 'MysticHammer', 
                   'ParryingDagger', 'PreciseHandCrossbow', 'ReliableMagicScepter', 'SwiftAxe', 
                   'TransferMagicOrb', 'WarpCrystal', 'GauntletOffhand', 'BoomerangOffhand'}

# Manual mapping for tricky weapon names
WEAPON_NAME_MAPPING = {
    'Axe': '1h Axe',
    'MagicWand': 'Wand',
    'MagicScepter': 'Scepter',
    'MagicStaff': 'power staff',
    'TomeOfMagic': 'Tome of Secrets',
    'DruidicStaff': 'druid staff',
    'WarShield': 'War Shield',
    '2HHammer': '2H Hammer',
    '2HAxe': '2H AXE',
    'HandCrossbow': 'Hand crossbow',
    'MagicOrb': 'Magic orb',
    'ManaFlower': 'Sacred Flower',
}

def find_excel_weapon_name(xml_base):
    """Find the Excel weapon name for a given XML weapon base (case-insensitive)"""
    # Check manual mapping first
    if xml_base in WEAPON_NAME_MAPPING:
        return WEAPON_NAME_MAPPING[xml_base]
    
    # Normalize the XML base name for comparison
    normalized = xml_base.lower()
    
    # Build case-insensitive lookup
    excel_weapons_lower = {k.lower(): k for k in weapon_variants_mapping.keys()}
    
    # Try direct lookup
    if normalized in excel_weapons_lower:
        return excel_weapons_lower[normalized]
    
    # Try removing hyphens and spaces from Excel names for comparison
    for excel_name, excel_name_lower in excel_weapons_lower.items():
        if excel_name_lower.replace(' ', '').replace('-', '') == normalized.replace(' ', '').replace('-', ''):
            return excel_name
    
    # Try checking if any excel weapon name contains the normalized string
    for excel_name, excel_name_lower in excel_weapons_lower.items():
        # Remove all non-alphanumeric for fuzzy matching
        excel_clean = ''.join(c for c in excel_name_lower if c.isalnum())
        xml_clean = ''.join(c for c in normalized if c.isalnum())
        if excel_clean == xml_clean:
            return excel_name
    
    return None

# ============================================================================
# PHASE 4: UPDATE WEAPON DAMAGE VALUES IN XML FILES
# ============================================================================
print("\n[PHASE 4] Updating weapon damage values...")
print("-" * 80)

def update_weapon_damage(file_path):
    """Process an ItemDefinitions XML file and update damage values"""
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    damage_changes = []
    update_count = 0
    
    for item in root.findall('.//ItemDefinition'):
        item_id = item.get('Id')
        if not item_id:
            continue
        
        # Extract variant ID from the weapon ID
        variant_id = int(item_id[-1]) if item_id[-1].isdigit() else None
        if variant_id is None:
            continue
        
        # Get the weapon base (strip off the trailing digit)
        weapon_base = item_id.rstrip('0123456789')
        
        # Check if this is an offhand weapon
        is_offhand = weapon_base in OFFHAND_WEAPONS
        
        # Find the Excel weapon name
        excel_weapon_name = find_excel_weapon_name(weapon_base)
        if not excel_weapon_name:
            continue
        
        # Get damage data for this weapon
        if excel_weapon_name not in weapon_data:
            continue
        
        levels_data = weapon_data[excel_weapon_name]['levels']
        
        # Process all level variants
        level_variations = item.find('LevelVariations')
        if level_variations is None:
            continue
        
        for level_elem in level_variations.findall('Level'):
            level_id_str = level_elem.get('Id')
            if level_id_str is None:
                continue
            
            try:
                level_id = int(level_id_str)
            except (ValueError, TypeError):
                continue
            
            # Map XML level to Excel level based on weapon type and variant_id
            # Non-offhand weapons ending in 0: use Excel levels -1 to 4 (XML 0→Excel -1, ..., XML 5→Excel 4)
            # All other weapons (including offhands): use Excel levels 0 to 5 (XML 0→Excel 0, ..., XML 5→Excel 5)
            if variant_id == 0 and not is_offhand:
                excel_level = level_id - 1  # XML 0→Excel -1, XML 1→Excel 0, ..., XML 5→Excel 4
            else:
                excel_level = level_id  # Direct mapping for offhands and all other weapons
            
            # Check if we have damage data for this Excel level
            if excel_level not in levels_data:
                continue
            
            damage_info = levels_data[excel_level]
            new_min = damage_info['min']
            new_max = damage_info['max']
            
            # Find or create BaseDamage element
            base_damage = level_elem.find('BaseDamage')
            if base_damage is None:
                base_damage = ET.SubElement(level_elem, 'BaseDamage')
            
            old_min = base_damage.get('Min')
            old_max = base_damage.get('Max')
            
            # Update damage values
            if old_min != str(new_min) or old_max != str(new_max):
                base_damage.set('Min', str(new_min))
                base_damage.set('Max', str(new_max))
                damage_changes.append({
                    'weapon_id': item_id,
                    'level': level_id,
                    'excel_level': excel_level,
                    'old': f"{old_min}-{old_max}" if old_min else "None",
                    'new': f"{new_min}-{new_max}"
                })
                update_count += 1
    
    # Write the updated XML
    tree.write(file_path, encoding='utf-8', xml_declaration=True)
    
    return update_count, damage_changes

# Process all weapon files for damage updates
all_damage_changes = []
total_damage_updates = 0

files_to_process_damage = [
    '../../modded_files/ItemDefinitions_Weapons',
    '../../modded_files/ItemDefinitions_DLC1',
    '../../modded_files/ItemDefinitions_DLC2',
]

for file_path in files_to_process_damage:
    print(f"\nProcessing {file_path}...")
    update_count, changes = update_weapon_damage(file_path)
    all_damage_changes.extend(changes)
    total_damage_updates += update_count
    print(f"  ✓ Updated {update_count} weapon damage values")

print(f"\n✓ Total weapon damage updates: {total_damage_updates}")

# ============================================================================
# PHASE 5: UPDATE WEAPON STAT BONUSES IN XML FILES
# ============================================================================
print("\n[PHASE 5] Updating weapon stat bonuses...")
print("-" * 80)

def parse_composite_value(value_str):
    """Parse composite values like '7;2' or '40;8' into a list"""
    if isinstance(value_str, str) and ';' in value_str:
        try:
            return [int(float(v.strip())) for v in value_str.split(';')]
        except:
            return []
    elif isinstance(value_str, (int, float)):
        return [int(value_str)]
    return []

def find_stat_value_in_bonuses(stat_name, bonuses_dict):
    """Find a stat value in bonuses dict, handling both simple and composite keys"""
    # First, try exact match
    if stat_name in bonuses_dict:
        return bonuses_dict[stat_name]
    
    # Try case-insensitive match
    for key, value in bonuses_dict.items():
        if key.lower() == stat_name.lower():
            return value
    
    # Try to find in composite keys (e.g., "Mana;Mana Regen")
    for key, value in bonuses_dict.items():
        if ';' in key:
            parts = [p.strip() for p in key.split(';')]
            for i, part in enumerate(parts):
                if part.lower() == stat_name.lower():
                    values = parse_composite_value(value)
                    if i < len(values):
                        return values[i]
    
    return None

def map_excel_stat_to_xml(excel_stat_name):
    """Map a single Excel stat name to XML stat name (case-insensitive)"""
    if excel_stat_name in stat_name_mapping:
        return stat_name_mapping[excel_stat_name]
    
    # Try case-insensitive match
    for key, val in stat_name_mapping.items():
        if key.lower() == excel_stat_name.lower():
            return val
    
    return None

def create_base_stat_bonuses(weapon_id, variant_id, level_id, excel_weapon_name):
    """Create BaseStatBonuses element for a weapon variant"""
    
    # Weapons ending in 0-1 should not have bonuses (except WarShield which always gets -20 Dodge)
    if variant_id == 0 or variant_id == 1:
        if excel_weapon_name.lower() == 'war shield':
            base_stat_bonuses = ET.Element('BaseStatBonuses')
            stat_elem = ET.SubElement(base_stat_bonuses, 'BaseStatBonus')
            stat_elem.set('Stat', 'Dodge')
            stat_elem.text = '-20'
            return base_stat_bonuses
        else:
            return None
    
    # Get the stat names for this variant
    variant_mapping = weapon_variants_mapping.get(excel_weapon_name)
    if not variant_mapping:
        return None
    
    # Variant keys can be int or str depending on how they were extracted
    if variant_id not in variant_mapping and str(variant_id) not in variant_mapping:
        return None
    
    stat_names = variant_mapping.get(variant_id) or variant_mapping.get(str(variant_id))
    
    # Normalize stat_names to always be a list
    if isinstance(stat_names, str):
        stat_names = [stat_names]
    
    # Determine which tier to use
    if variant_id in [2, 3]:
        bonuses_dict = tier1_bonuses.get(level_id, {})
    elif variant_id in [4, 5]:
        bonuses_dict = tier2_bonuses.get(level_id, {})
    else:
        return None
    
    # Create BaseStatBonuses element
    base_stat_bonuses = ET.Element('BaseStatBonuses')
    
    for excel_stat_name in stat_names:
        excel_stat_name = excel_stat_name.strip()
        
        # Map to XML stat name
        xml_stat_name = map_excel_stat_to_xml(excel_stat_name)
        if not xml_stat_name:
            continue
        
        # Find the value in bonuses dict
        value = find_stat_value_in_bonuses(excel_stat_name, bonuses_dict)
        if value is None:
            continue
        
        # Handle composite values and map to correct XML stat
        if isinstance(xml_stat_name, list):
            # Multi-stat mapping (e.g., ["MovePointsTotal", "Dodge"])
            values = parse_composite_value(value)
            for i, stat in enumerate(xml_stat_name):
                if i < len(values):
                    stat_elem = ET.SubElement(base_stat_bonuses, 'BaseStatBonus')
                    stat_elem.set('Stat', stat)
                    stat_elem.text = str(values[i])
        else:
            # Single stat mapping
            values = parse_composite_value(value)
            if values:
                stat_elem = ET.SubElement(base_stat_bonuses, 'BaseStatBonus')
                stat_elem.set('Stat', xml_stat_name)
                stat_elem.text = str(values[0])
    
    # For WarShield, always add -20 Dodge
    if excel_weapon_name.lower() == 'war shield':
        dodge_exists = any(elem.get('Stat') == 'Dodge' for elem in base_stat_bonuses.findall('BaseStatBonus'))
        if not dodge_exists:
            stat_elem = ET.SubElement(base_stat_bonuses, 'BaseStatBonus')
            stat_elem.set('Stat', 'Dodge')
            stat_elem.text = '-20'
        else:
            for elem in base_stat_bonuses.findall('BaseStatBonus'):
                if elem.get('Stat') == 'Dodge':
                    elem.text = '-20'
    
    return base_stat_bonuses if len(base_stat_bonuses) > 0 else None

def process_xml_file(file_path):
    """Process an ItemDefinitions XML file and update stat bonuses"""
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    changes = []
    update_count = 0
    skipped_weapons = []
    
    for item in root.findall('.//ItemDefinition'):
        item_id = item.get('Id')
        if not item_id:
            continue
        
        # Extract variant ID from the weapon ID
        variant_id = int(item_id[-1]) if item_id[-1].isdigit() else None
        if variant_id is None:
            continue
        
        # Get the weapon base (strip off the trailing digit)
        weapon_base = item_id.rstrip('0123456789')
        
        # Skip OffHand weapons
        if weapon_base in OFFHAND_WEAPONS:
            continue
        
        # Find the Excel weapon name (case-insensitive)
        excel_weapon_name = find_excel_weapon_name(weapon_base)
        if not excel_weapon_name:
            if weapon_base not in [w[0] for w in skipped_weapons]:
                skipped_weapons.append((weapon_base, item_id))
            continue
        
        # Process all level variants
        level_variations = item.find('LevelVariations')
        if level_variations is None:
            continue
        
        for level_elem in level_variations.findall('Level'):
            level_id_str = level_elem.get('Id')
            if level_id_str is None:
                continue
            
            try:
                level_id = int(level_id_str)
            except (ValueError, TypeError):
                continue
            
            # Create new BaseStatBonuses
            new_base_stat_bonuses = create_base_stat_bonuses(item_id, variant_id, level_id, excel_weapon_name)
            
            # Remove old BaseStatBonuses if it exists
            old_bsb = level_elem.find('BaseStatBonuses')
            if old_bsb is not None:
                level_elem.remove(old_bsb)
            
            # Add new BaseStatBonuses
            if new_base_stat_bonuses is not None:
                level_elem.append(new_base_stat_bonuses)
                
                # Log the change
                bonus_str = ', '.join([
                    f"{stat.get('Stat')}={stat.text}"
                    for stat in new_base_stat_bonuses.findall('BaseStatBonus')
                ])
                changes.append({
                    'weapon_id': item_id,
                    'variant_id': variant_id,
                    'level': level_id,
                    'bonuses': bonus_str
                })
                update_count += 1
    
    # Validate tree has content before writing
    item_count = len(root.findall('.//ItemDefinition'))
    if item_count == 0:
        print(f"  ⚠ ERROR: Tree is empty, aborting write for {file_path}")
        return update_count, changes
    
    original_size = os.path.getsize(file_path)
    
    # Write the updated XML
    tree.write(file_path, encoding='utf-8', xml_declaration=True)
    
    # Verify file was written successfully
    new_size = os.path.getsize(file_path)
    if new_size < original_size * 0.5:  # File shrank by more than 50%
        print(f"  ⚠ ERROR: File size dropped from {original_size} to {new_size} bytes - aborting")
        return update_count, changes
    
    print(f"  ✓ File saved: {item_count} items, {new_size} bytes")
    
    # Print skipped weapons for debugging
    if skipped_weapons:
        print(f"\n  DEBUG: Skipped {len(skipped_weapons)} weapon bases (no Excel match):")
        for weapon_base, example_id in skipped_weapons[:10]:
            print(f"    {weapon_base} (e.g., {example_id})")
    
    return update_count, changes

# Process all weapon files
all_stat_changes = []
total_stat_updates = 0

files_to_process = [
    '../../modded_files/ItemDefinitions_Weapons',
    '../../modded_files/ItemDefinitions_DLC1',
    '../../modded_files/ItemDefinitions_DLC2',
]

for file_path in files_to_process:
    print(f"\nProcessing {file_path}...")
    update_count, changes = process_xml_file(file_path)
    all_stat_changes.extend(changes)
    total_stat_updates += update_count
    print(f"  ✓ Updated {update_count} BaseStatBonuses")

print(f"\n✓ Total stat bonus updates: {total_stat_updates}")

# ============================================================================
# PHASE 6: UPDATE SCROLL ITEM DAMAGE VALUES IN XML FILES
# ============================================================================
print("\n[PHASE 6] Updating scroll item damage values...")
print("-" * 80)

scroll_mapping = {
    'AxeBoomerangScroll': ('1h Axe', 'Axe'),
    'ThrowingDaggersScroll': ('Dagger', 'Dagger'),
    'ChargeScroll': ('2h sword', '2HSword'),
    'SwordBlastScroll': ('2h sword', '2HSword'),
    'SuperSpinScroll': ('2H AXE', '2HAxe'),
    'GroundSmashScroll': ('2H Hammer', '2HHammer'),
    'TripleSwipeScroll': ('Spear', 'Spear'),
    'GrapeshotScroll': ('Pistol', 'Pistol'),
    'RainOfArrowsScroll': ('Shortbow', 'Shortbow'),
    'ExplosiveBoltScroll': ('Crossbow', 'Crossbow'),
    'AssassinateScroll': ('Rifle', 'Rifle'),
    'MagicMissilesScroll': ('Wand', 'Wand'),
    'HammerOfFaithScroll': ('Scepter', 'Scepter'),
    'DeathRayScroll': ('Magic orb', 'Magic orb'),
    'ScorchingWaveScroll': ('power staff', 'power staff'),
    'FireThrowerScroll': ('power staff', 'power staff'),
    'FireballScroll': ('Tome of Secrets', 'Tome of Secrets'),
    'LightningStrikeScroll': ('Tome of Secrets', 'Tome of Secrets'),
    'BeeStingScroll': ('druid staff', 'druid staff'),
    'TeleportationScroll': None,
}

tree = ET.parse('../../modded_files/ItemDefinitions_Usables')
root = tree.getroot()

scroll_changes = []
total_scroll_updated = 0
total_removed = 0

for item_def in root.findall('ItemDefinition'):
    item_id = item_def.get('Id')
    
    if item_id not in scroll_mapping:
        continue
    
    mapping = scroll_mapping[item_id]
    
    # Special case: TeleportationScroll should not have BaseDamage elements
    if mapping is None:
        print(f"\n{item_id}: removing damage values")
        level_variations = item_def.find('LevelVariations')
        if level_variations is not None:
            for level_elem in level_variations.findall('Level'):
                base_damage = level_elem.find('BaseDamage')
                if base_damage is not None:
                    level_id = level_elem.get('Id')
                    level_elem.remove(base_damage)
                    total_removed += 1
        continue
    
    excel_sheet, weapon_prefix = mapping
    
    if excel_sheet not in weapon_data:
        print(f"Warning: {excel_sheet} not in weapon data")
        continue
    
    levels = weapon_data[excel_sheet]['levels']
    print(f"\n{item_id} (from {excel_sheet}):")
    
    level_mapping = {0: 0, 1: 1, 2: 2, 3: 3, 4: 4, 5: 5}
    
    level_variations = item_def.find('LevelVariations')
    if level_variations is None:
        continue
    
    updated_count = 0
    for level_elem in level_variations.findall('Level'):
        level_id = level_elem.get('Id')
        try:
            level_id_int = int(level_id)
        except (ValueError, TypeError):
            continue
        
        if level_id_int not in level_mapping:
            continue
        
        excel_level = level_mapping[level_id_int]
        excel_level_str = str(excel_level)
        
        if excel_level_str not in levels:
            continue
        
        base_damage = level_elem.find('BaseDamage')
        if base_damage is None:
            continue
        
        old_min = base_damage.get('Min')
        old_max = base_damage.get('Max')
        new_min = levels[excel_level_str]['min']
        new_max = levels[excel_level_str]['max']
        
        if old_min != str(new_min) or old_max != str(new_max):
            base_damage.set('Min', str(new_min))
            base_damage.set('Max', str(new_max))
            scroll_changes.append({
                'scroll': item_id,
                'level': level_id,
                'old': f"{old_min}-{old_max}",
                'new': f"{new_min}-{new_max}"
            })
            updated_count += 1
            total_scroll_updated += 1
            print(f"  Level {level_id}: {old_min}-{old_max} -> {new_min}-{new_max}")
    
    if updated_count == 0:
        print(f"  No updates needed")

tree.write('../../modded_files/ItemDefinitions_Usables', encoding='utf-8', xml_declaration=True)
print(f"\nUpdated {total_scroll_updated} scroll damage values")
print(f"Removed {total_removed} damage values from special scrolls")

# ============================================================================
# PHASE 7: REFORMAT ALL XML FILES
# ============================================================================
print("\n[PHASE 7] Reformatting XML files...")
print("-" * 80)

def indent_xml(elem, level=0):
    """Add proper indentation to XML elements"""
    indent = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent

def reformat_xml_file(file_path):
    """Reformat an XML file with proper indentation and spacing"""
    tree = ET.parse(file_path)
    root = tree.getroot()
    indent_xml(root)
    
    # Add blank line after each ItemDefinition
    for item_def in root.findall('ItemDefinition'):
        if item_def.tail and item_def.tail.strip() == '':
            # Add extra newline for spacing between ItemDefinitions
            item_def.tail = '\n\n' + (root.tag == item_def.tag and '' or '  ')
    
    tree.write(file_path, encoding='utf-8', xml_declaration=True)

# Reformat all modified files
files_to_reformat = [
    '../../modded_files/ItemDefinitions_Weapons',
    '../../modded_files/ItemDefinitions_DLC1',
    '../../modded_files/ItemDefinitions_DLC2',
    '../../modded_files/ItemDefinitions_Usables',
]

for file_path in files_to_reformat:
    try:
        reformat_xml_file(file_path)
        print(f"Reformatted {file_path}")
    except Exception as e:
        print(f"Error reformatting {file_path}: {e}")

# ============================================================================
# FINAL SUMMARY
# ============================================================================
print("\n" + "=" * 80)
print("CONSOLIDATION COMPLETE")
print("=" * 80)
print(f"Weapon damage updates: {total_damage_updates}")
print(f"Stat bonus updates: {total_stat_updates}")
print(f"Scroll damage updates: {total_scroll_updated}")
print(f"Scroll removals: {total_removed}")
print(f"All XML files have been updated, synchronized, and reformatted")
print("=" * 80)
